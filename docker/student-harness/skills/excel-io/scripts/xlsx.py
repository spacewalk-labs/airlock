#!/usr/bin/env -S uv run --quiet --script
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Excel workbook survey, extraction, verification and recalculation.

    probe    structure first: sheets, used range, headers, types, formulas
    dump     one sheet (or range) as a markdown or TSV table
    check    output gate: error values, broken refs, missing cached values
    recalc   let LibreOffice compute formulas openpyxl only wrote as text

probe before dump. A workbook's used range is routinely 100x larger than the
data in it — stray formatting stretches `max_row` into the thousands — so the
first read should report shape, not content.

probe/dump/check never modify the workbook. recalc writes, and says where.
"""

from __future__ import annotations

import argparse
import datetime as dt
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

ERROR_VALUES = ("#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#SPILL!")


def die(msg: str) -> "None":
    sys.exit(f"error: {msg}")


def check_path(path_str: str) -> Path:
    path = Path(path_str).expanduser()
    if not path.is_file():
        die(f"not a file: {path}")
    suffix = path.suffix.lower()
    if suffix == ".xls":
        die(
            "this is a legacy .xls (BIFF) file, which openpyxl cannot read.\n"
            f"  convert first: soffice --headless --convert-to xlsx --outdir /tmp '{path}'"
        )
    if suffix == ".csv":
        die("this is a CSV — read it directly, no Excel tooling needed.")
    if suffix not in (".xlsx", ".xlsm"):
        die(f"unsupported extension '{suffix}'. This tool reads .xlsx and .xlsm.")
    return path


def fmt(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()[:19]
    return str(value)


def profile(value) -> str:
    if value is None:
        return "blank"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, (dt.datetime, dt.date)):
        return "date"
    if isinstance(value, (int, float)):
        return "number"
    text = str(value)
    if text.startswith("="):
        return "formula"
    # A number typed with a thousands separator or a trailing space arrives as
    # text and silently drops out of every SUM downstream.
    stripped = text.strip()
    if stripped.startswith("(") and stripped.endswith(")"):
        stripped = "-" + stripped[1:-1]          # (1,200) = accounting negative
    for junk in (",", " ", "%", "\u20a9", "$", "\u20ac", "\u00a5", "\u00a3", "\xa0"):
        stripped = stripped.replace(junk, "")
    try:
        float(stripped)
        return "number-as-text"
    except ValueError:
        return "text"


# --------------------------------------------------------------------------


def cmd_probe(args) -> None:
    path = check_path(args.file)
    wb = openpyxl.load_workbook(path, data_only=False)
    wbv = openpyxl.load_workbook(path, data_only=True)
    print(f"file    {path}")
    print(f"sheets  {len(wb.sheetnames)}: {', '.join(wb.sheetnames)}")

    for name in wb.sheetnames:
        ws = wb[name]
        wsv = wbv[name]
        formulas = cached = 0
        types: dict[str, int] = {}
        last_row = last_col = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                last_row = max(last_row, cell.row)
                last_col = max(last_col, cell.column)
                kind = profile(cell.value)
                types[kind] = types.get(kind, 0) + 1
                if kind == "formula":
                    formulas += 1
                    if wsv[cell.coordinate].value is not None:
                        cached += 1

        used = f"A1:{get_column_letter(last_col)}{last_row}" if last_row else "(empty)"
        declared = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        print(f"\n[{name}]  state={ws.sheet_state}")
        print(f"  used range      {used}"
              + (f"    (declared {declared} — formatting-only padding)"
                 if used != declared and last_row else ""))
        print(f"  cell types      {dict(sorted(types.items(), key=lambda kv: -kv[1]))}")
        if formulas:
            note = "all cached" if cached == formulas else f"{formulas - cached} WITHOUT cached value"
            print(f"  formulas        {formulas} ({note})")
        if ws.merged_cells.ranges:
            shown = [str(r) for r in list(ws.merged_cells.ranges)[:8]]
            print(f"  merged ranges   {len(ws.merged_cells.ranges)}: {shown}"
                  "  <- only the top-left cell holds the value")
        if ws.freeze_panes:
            print(f"  freeze panes    {ws.freeze_panes}")
        if last_row:
            header_row, header = _guess_header(ws, last_col)
            print(f"  header guess    row {header_row}: {header}")
            for r in ws.iter_rows(min_row=header_row + 1, max_row=min(header_row + args.preview,
                                                                     last_row),
                                  max_col=last_col, values_only=True):
                print("    | " + " | ".join(fmt(v) for v in r))


def _guess_header(ws, last_col: int):
    """First row whose cells are mostly non-blank text — the usual header shape.

    Report the row number rather than assuming row 1: Korean workbooks routinely
    carry a merged title banner and a blank spacer above the real header.
    """
    for row in ws.iter_rows(min_row=1, max_row=min(50, ws.max_row), max_col=last_col):
        values = [c.value for c in row]
        filled = [v for v in values if v is not None]
        if len(filled) >= max(2, len(values) * 0.5) and all(
            isinstance(v, str) for v in filled
        ):
            return row[0].row, [fmt(v) for v in values]
    first = next(ws.iter_rows(min_row=1, max_row=1, max_col=last_col, values_only=True), ())
    return 1, [fmt(v) for v in first]


def cmd_dump(args) -> None:
    path = check_path(args.file)
    wb = openpyxl.load_workbook(path, data_only=not args.formulas, read_only=True)
    name = args.sheet or wb.sheetnames[0]
    if name not in wb.sheetnames:
        sheets = ", ".join(wb.sheetnames)
        wb.close()
        die(f"no sheet named {name!r}. Sheets: {sheets}")
    ws = wb[name]
    # ws["A1"] yields a bare cell, ws["A1:B2"] a tuple of tuples, and a whole
    # column a flat tuple. Normalise all three to rows-of-cells.
    if args.range:
        try:
            selection = ws[args.range]
        except AttributeError:
            # An unbounded column range ("A:B") needs iter_cols, which the
            # streaming reader does not have. Reopen the normal way: the range
            # the caller asked for is what they get, not a traceback.
            wb.close()
            wb = openpyxl.load_workbook(path, data_only=not args.formulas)
            ws = wb[name]
            selection = ws[args.range]
        if not isinstance(selection, tuple):
            rows = [(selection,)]
        elif selection and not isinstance(selection[0], tuple):
            rows = [(cell,) for cell in selection]
        else:
            rows = list(selection)
            # "A:B" hands back a tuple per *column*, while "A1:B3" hands back one
            # per row. Both are tuples of tuples, so printing blindly transposes
            # the table — wrong output that still looks like a table. Tell them
            # apart by whether the first group shares one column.
            first = rows[0] if rows else ()
            if len(first) > 1 and len({c.column for c in first}) == 1:
                rows = [tuple(group) for group in zip(*rows)]
    else:
        # Stop pulling rows at the limit instead of materialising the sheet:
        # read_only mode streams, and a padded sheet can be a million rows.
        rows = []
        for row in ws.iter_rows():
            rows.append(row)
            if len(rows) >= args.max_rows:
                break

    table = [[fmt(c.value) for c in row] for row in rows][: args.max_rows]
    while table and not any(cell for cell in table[-1]):
        table.pop()
    wb.close()
    if not table:
        print(f"# sheet {name!r} is empty in the requested range")
        return

    width = max(len(r) for r in table)
    table = [r + [""] * (width - len(r)) for r in table]
    print(f"# sheet {name!r} · {len(table)} rows x {width} cols"
          + (f" · range {args.range}" if args.range else "")
          + (" · formulas" if args.formulas else " · values"))
    if args.format == "tsv":
        for r in table:
            print("\t".join(r))
        return
    header, *body = table
    print("| " + " | ".join(header) + " |")
    print("|" + "|".join(["---"] * width) + "|")
    for r in body:
        print("| " + " | ".join(c.replace("|", "\\|") for c in r) + " |")


def cmd_check(args) -> None:
    path = check_path(args.file)
    wbf = openpyxl.load_workbook(path, data_only=False)
    wbv = openpyxl.load_workbook(path, data_only=True)
    problems: list[str] = []
    formulas = uncached = 0

    for name in wbf.sheetnames:
        wsf, wsv = wbf[name], wbv[name]
        for row in wsf.iter_rows():
            for cell in row:
                text = str(cell.value) if cell.value is not None else ""
                if text.startswith("="):
                    formulas += 1
                    # Only #REF! is scanned in the formula text. The other codes
                    # appear legitimately as string literals — =IFERROR(x,"#N/A")
                    # is correct code, and flagging it trains people to ignore
                    # this gate. A deleted reference has no such innocent form.
                    if "#REF!" in text:
                        problems.append(f"{name}!{cell.coordinate}: formula references #REF!"
                                        f"  {text[:80]}")
                    if wsv[cell.coordinate].value is None:
                        uncached += 1
                    cached = wsv[cell.coordinate].value
                    if isinstance(cached, str) and cached in ERROR_VALUES:
                        problems.append(f"{name}!{cell.coordinate}: formula evaluates to {cached}"
                                        f"  {text[:80]}")
                elif text in ERROR_VALUES:
                    problems.append(f"{name}!{cell.coordinate}: stored error value {text}")

    print(f"checked {path} · {len(wbf.sheetnames)} sheets · {formulas} formulas")
    if uncached:
        print(f"WARN  {uncached} of {formulas} formulas have no cached value.\n"
              "      Anything reading this file with data_only=True sees None there.\n"
              f"      Fix: {Path(__file__).name} recalc '{path}'")
    if problems:
        print(f"FAIL  {len(problems)} error cells")
        for line in problems[:50]:
            print(f"  {line}")
        if len(problems) > 50:
            print(f"  ... {len(problems) - 50} more")
        sys.exit(1)
    print("PASS  no #REF!/#DIV0!/#VALUE!/#N/A/#NAME?/#NULL!/#NUM! anywhere")
    if uncached:
        sys.exit(2)


def cmd_recalc(args) -> None:
    """openpyxl writes formula text but never computes it, so a workbook it
    produced has no cached values. LibreOffice recalculates on load and stores
    the results, which is what makes the file readable by value consumers."""
    path = check_path(args.file)
    exe = shutil.which("soffice") or shutil.which("libreoffice")
    if not exe:
        die("LibreOffice is not installed (needed to compute formulas).\n"
            "  install: sudo apt-get install -y libreoffice-calc")
    if args.output:
        out = Path(args.output).expanduser()
    elif args.in_place:
        out = path
    else:
        # Default to a sibling file. Recalculating means handing the workbook to
        # a different program and writing back whatever it produces; making that
        # overwrite the only copy by default is a poor trade for one saved flag.
        out = path.with_suffix(f".recalc{path.suffix}")
    with tempfile.TemporaryDirectory(prefix="xlsx-recalc-") as tmp:
        proc = subprocess.run(
            [exe, "--headless", "--norestore", "--convert-to", "xlsx", "--outdir", tmp, str(path)],
            capture_output=True, text=True, timeout=300,
        )
        produced = sorted(Path(tmp).glob("*.xlsx"))
        if not produced:
            die(f"LibreOffice produced nothing (exit {proc.returncode}): "
                f"{proc.stderr.strip()[:300] or proc.stdout.strip()[:300]}")
        shutil.copy(produced[0], out)

    wbv = openpyxl.load_workbook(out, data_only=True)
    wbf = openpyxl.load_workbook(out, data_only=False)
    total = filled = 0
    for name in wbf.sheetnames:
        for row in wbf[name].iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    total += 1
                    if wbv[name][cell.coordinate].value is not None:
                        filled += 1
    print(f"recalculated -> {out}")
    print(f"  {filled}/{total} formulas now carry a cached value")
    if total and filled < total:
        print("  the remainder are formulas LibreOffice could not evaluate — inspect them")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    sub = parser.add_subparsers(dest="cmd", required=True)

    p = sub.add_parser("probe", help="sheets, used range, headers, types, formulas")
    p.add_argument("file")
    p.add_argument("--preview", type=int, default=3, help="data rows to show per sheet")
    p.set_defaults(func=cmd_probe)

    p = sub.add_parser("dump", help="a sheet or range as a markdown/TSV table")
    p.add_argument("file")
    p.add_argument("--sheet")
    p.add_argument("--range", help="e.g. A1:H40")
    p.add_argument("--max-rows", type=int, default=200)
    p.add_argument("--format", choices=("md", "tsv"), default="md")
    p.add_argument("--formulas", action="store_true", help="show formula text instead of values")
    p.set_defaults(func=cmd_dump)

    p = sub.add_parser("check", help="error-value gate; exits non-zero when it fails")
    p.add_argument("file")
    p.set_defaults(func=cmd_check)

    p = sub.add_parser("recalc", help="compute formulas via LibreOffice and store the results")
    p.add_argument("file")
    p.add_argument("-o", "--output", help="write here (default: <name>.recalc.xlsx)")
    p.add_argument("--in-place", action="store_true",
                   help="overwrite the input file instead of writing a sibling")
    p.set_defaults(func=cmd_recalc)

    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
